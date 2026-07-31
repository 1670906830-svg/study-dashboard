import os
import re
from datetime import datetime, date, timedelta
from pathlib import Path
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession, async_sessionmaker
from sqlalchemy import text

# ---- Database URL resolution ----
_raw_url = os.environ.get("DATABASE_URL", "")

if _raw_url:
    if _raw_url.startswith("postgres://"):
        _db_url = _raw_url.replace("postgres://", "postgresql+asyncpg://", 1)
    elif _raw_url.startswith("postgresql://"):
        _db_url = _raw_url.replace("postgresql://", "postgresql+asyncpg://", 1)
    else:
        _db_url = _raw_url
else:
    _db_url = "sqlite+aiosqlite:///study.db"

engine = create_async_engine(_db_url, echo=False, future=True)
async_session = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)


def is_postgres() -> bool:
    return "postgresql" in _db_url or "+asyncpg" in _db_url


def row_to_dict(row):
    if row is None:
        return None
    return dict(row._mapping)


async def get_db():
    async with async_session() as session:
        yield session


# ---- DDL (dialect-aware) ----

async def init_db():
    pg = is_postgres()
    pk_type = "SERIAL" if pg else "INTEGER"
    pk_extra = " PRIMARY KEY" if pg else " PRIMARY KEY AUTOINCREMENT"
    ref = "REFERENCES" if pg else "REFERENCES"

    async with engine.begin() as conn:
        await conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS courses (
                id {pk_type}{pk_extra},
                name TEXT NOT NULL,
                short_name TEXT NOT NULL,
                category TEXT NOT NULL CHECK(category IN ('行测', '申论', '健身')),
                total_sessions INTEGER DEFAULT 0,
                completed_sessions INTEGER DEFAULT 0,
                sort_order INTEGER DEFAULT 0
            );
        """))

        await conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS schedule (
                id {pk_type}{pk_extra},
                course_id INTEGER NOT NULL {ref} courses(id),
                date TEXT NOT NULL,
                session_label TEXT DEFAULT '',
                completed INTEGER DEFAULT 0
            );
        """))

        await conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS habits (
                id {pk_type}{pk_extra},
                name TEXT NOT NULL UNIQUE,
                frequency TEXT NOT NULL CHECK(frequency IN ('daily', 'alternate')) DEFAULT 'daily',
                sort_order INTEGER DEFAULT 0
            );
        """))

        await conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS habit_logs (
                id {pk_type}{pk_extra},
                habit_id INTEGER NOT NULL {ref} habits(id),
                date TEXT NOT NULL,
                completed INTEGER DEFAULT 0,
                completed_at TEXT,
                UNIQUE(habit_id, date)
            );
        """))

        await conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS correct_rates (
                id {pk_type}{pk_extra},
                category TEXT NOT NULL CHECK(category IN ('行测', '申论', '健身')),
                date TEXT NOT NULL,
                rate REAL NOT NULL CHECK(rate >= 0 AND rate <= 100),
                note TEXT DEFAULT '',
                UNIQUE(category, date)
            );
        """))

        for idx_sql in [
            "CREATE INDEX IF NOT EXISTS idx_schedule_date ON schedule(date)",
            "CREATE INDEX IF NOT EXISTS idx_schedule_course ON schedule(course_id)",
            "CREATE INDEX IF NOT EXISTS idx_habit_logs_date ON habit_logs(date)",
            "CREATE INDEX IF NOT EXISTS idx_correct_rates ON correct_rates(category, date)",
        ]:
            await conn.execute(text(idx_sql))


# ====================================================================
#  Data-import utilities (local-only, uses synchronous SQLite)
# ====================================================================

def _sqlite_connect():
    import sqlite3
    conn = sqlite3.connect(str(Path(__file__).parent / "study.db"))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def normalize_course_name(raw: str) -> str:
    s = raw.strip()
    s = re.sub(r'[「【][^」】]+[」】]', '', s).strip()
    s = re.sub(r'[（(]\d+\.?\d*H[)）]', '', s).strip()
    s = re.sub(r'\d+$', '', s).strip()
    return s


def find_course_match(raw_name: str, course_map: dict) -> str:
    normalized = normalize_course_name(raw_name)
    if normalized in course_map:
        return normalized
    explicit = {
        '粉笔-判断': '粉笔-判断',
        '冯泽-政治理论': '冯泽-政治理论',
        '花生-资料分析': '花生-资料分析',
        '花生-言语-片段': '花生-言语-片段-逻辑-总结',
        '花生-言语-逻辑': '花生-言语-片段-逻辑-总结',
        '花生-言语-总结': '花生-言语-片段-逻辑-总结',
        '花生--逻辑': '花生-判断-定义-逻辑-图推-总结',
        '花生-判断-逻辑': '花生-判断-定义-逻辑-图推-总结',
        '花生-判断-图推': '花生-判断-定义-逻辑-图推-总结',
        '花生-判断-定义': '花生-判断-定义-逻辑-图推-总结',
        '花生-判断-总结': '花生-判断-定义-逻辑-图推-总结',
        '王君涛-政治': '王君涛-政治',
        '花生--数量关系': '花生--数量关系',
        '四海-行测精讲': '四海-行测精讲',
        '粉笔-常识': '粉笔-常识',
        '王君涛-常识': '王君涛-常识',
        '飞扬-申论': '飞扬-申论-总结',
        '飞扬-申论-总结': '飞扬-申论-总结',
        '粉笔-单一题': '粉笔-单一题-综合题',
        '粉笔-综合题': '粉笔-单一题-综合题',
        '粉笔-公文题': '粉笔-公文题-材料作文',
        '粉笔-材料作文': '粉笔-公文题-材料作文',
        '言语-1000词': '言语-1000词',
        '龙飞-图推': '龙飞-图推',
        '江苏': '江苏',
        '数推刷题': '数推刷题',
        '资料刷题': '资料刷题',
        '资料套题': '资料刷题',
        '图推刷题': '判断推理刷题',
        '定义刷题': '判断推理刷题',
        '类比刷题': '判断推理刷题',
        '逻辑刷题': '判断推理刷题',
        '判断套题': '判断推理刷题',
        '片段刷题': '言语刷题',
        '选词刷题': '言语刷题',
        '言语套题': '言语刷题',
        '数量刷题': '数量刷题',
        '数量综合刷题': '数量刷题',
        '粉笔-常识-中古史': '粉笔-常识',
        '粉笔-常识-中现史': '粉笔-常识',
        '粉笔-常识-世界史': '粉笔-常识',
        '粉笔-常识-中传': '粉笔-常识',
        '粉笔-常识-中文常': '粉笔-常识',
        '粉笔-常识-世文化': '粉笔-常识',
        '粉笔-常识-生物': '粉笔-常识',
        '粉笔-常识-化学': '粉笔-常识',
        '粉笔-常识-科技': '粉笔-常识',
        '粉笔-常识-物理': '粉笔-常识',
        '粉笔-常识-地理': '粉笔-常识',
        '粉笔-常识-中地': '粉笔-常识',
        '粉笔-常识-经济': '粉笔-常识',
        '粉笔-常识-宪法': '粉笔-常识',
        '粉笔-常识-行政法': '粉笔-常识',
        '粉笔-常识-民法': '粉笔-常识',
        '粉笔-常识-刑法': '粉笔-常识',
        '粉笔-常识-科技': '粉笔-常识',
        '江苏-主观题基础': '江苏',
        '江苏-概括做法': '江苏',
        '江苏-提出对策': '江苏',
        '江苏-综合题': '江苏',
        '江苏-综合对比分析': '江苏',
        '江苏-公文底层逻辑': '江苏',
        '江苏-总结类': '江苏',
        '江苏-权威类': '江苏',
        '江苏-方案+启示类': '江苏',
        '江苏-材料+文章类': '江苏',
        '江苏-高分认知': '江苏',
        '江苏-主题与观点': '江苏',
        '江苏-落笔成文': '江苏',
    }
    for key, val in explicit.items():
        if key in normalized or normalized.startswith(key):
            return val
    best = ''
    for cname in course_map:
        if normalized.startswith(cname) and len(cname) > len(best):
            best = cname
        if cname.startswith(normalized) and len(cname) > len(best):
            best = cname
        if cname in normalized and len(cname) > len(best):
            best = cname
        if normalized in cname and len(cname) > len(best):
            best = cname
    if best:
        return best
    return normalized


def import_data(schedule_path: str, course_path: str):
    import openpyxl
    conn = _sqlite_connect()
    cur = conn.cursor()

    cur.execute("DELETE FROM schedule")
    cur.execute("DELETE FROM courses")
    cur.execute("DELETE FROM habit_logs")
    cur.execute("DELETE FROM habits")
    cur.execute("DELETE FROM correct_rates")

    wb_course = openpyxl.load_workbook(course_path)
    ws_course = wb_course['Sheet1']
    course_map = {}
    sort_idx = 0
    current_cat = '行测'
    for row in ws_course.iter_rows(min_row=2, values_only=True):
        cat = row[0]
        if cat and str(cat).strip() in ('行测', '申论'):
            current_cat = str(cat).strip()
        name = str(row[1]).strip() if row[1] else ''
        total = int(row[2]) if row[2] else 0
        if not name:
            continue
        sort_idx += 1
        cur.execute(
            "INSERT INTO courses (name, short_name, category, total_sessions, completed_sessions, sort_order) VALUES (?, ?, ?, ?, 0, ?)",
            (name, name, current_cat, total, sort_idx),
        )
        course_id = cur.lastrowid
        course_map[name] = (course_id, total, current_cat)

    wb_sch = openpyxl.load_workbook(schedule_path)
    ws_sch = wb_sch['Sheet1']

    def excel_serial_to_date(serial):
        base = datetime(1899, 12, 30)
        return base + timedelta(days=serial)

    date_rows_positions = []
    for r in range(3, ws_sch.max_row + 1):
        v = ws_sch.cell(row=r, column=1).value
        if isinstance(v, (int, float)) and 46000 < v < 47000:
            date_rows_positions.append(r)

    unmatched = []
    for pos in date_rows_positions:
        date_vals = [ws_sch.cell(row=pos, column=c).value for c in range(1, 6)]
        for offset in range(1, 6):
            check_pos = pos + offset
            if check_pos > ws_sch.max_row:
                break
            first_cell = ws_sch.cell(row=check_pos, column=1).value
            if isinstance(first_cell, (int, float)) and 46000 < first_cell < 47000:
                break
            row_cells = [ws_sch.cell(row=check_pos, column=c).value for c in range(1, 6)]
            if not any(c is not None for c in row_cells):
                continue
            for col in range(5):
                if date_vals[col] is not None and row_cells[col] is not None:
                    dt = excel_serial_to_date(date_vals[col]) + timedelta(days=1)
                    date_str = dt.strftime("%Y-%m-%d")
                    raw_name = str(row_cells[col]).strip()
                    matched = find_course_match(raw_name, course_map)
                    if matched and matched in course_map:
                        course_id, _, _ = course_map[matched]
                        cur.execute(
                            "INSERT INTO schedule (course_id, date, session_label) VALUES (?, ?, ?)",
                            (course_id, date_str, raw_name),
                        )
                    else:
                        unmatched.append(raw_name)

    if unmatched:
        print(f"Warning: {len(unmatched)} unmatched schedule items:")
        for u in unmatched[:5]:
            print(f"  {u}")

    habit_names = ["晨读", "练字", "喝水", "新闻联播", "健身"]
    habit_freqs = ["daily", "daily", "daily", "daily", "alternate"]
    for i, (hname, hfreq) in enumerate(zip(habit_names, habit_freqs), 1):
        cur.execute(
            "INSERT INTO habits (name, frequency, sort_order) VALUES (?, ?, ?)",
            (hname, hfreq, i),
        )
        cur.execute("SELECT id FROM habits WHERE name = ?", (hname,))
        hid = cur.fetchone()["id"]
        start = date(2026, 7, 31)
        end = date(2026, 11, 28)
        current = start
        while current <= end:
            date_str = current.strftime("%Y-%m-%d")
            if hfreq == "daily":
                cur.execute(
                    "INSERT OR IGNORE INTO habit_logs (habit_id, date) VALUES (?, ?)",
                    (hid, date_str),
                )
            elif hfreq == "alternate":
                days_since = (current - start).days
                if days_since % 2 == 0:
                    cur.execute(
                        "INSERT OR IGNORE INTO habit_logs (habit_id, date) VALUES (?, ?)",
                        (hid, date_str),
                    )
            current += timedelta(days=1)

    conn.commit()
    cur.execute("SELECT COUNT(*) as c FROM courses")
    total_courses = cur.fetchone()["c"]
    cur.execute("SELECT COUNT(*) as c FROM schedule")
    total_schedule = cur.fetchone()["c"]
    cur.execute("SELECT COUNT(*) as c FROM habit_logs")
    total_habits = cur.fetchone()["c"]
    conn.close()
    print(f"Courses: {total_courses}")
    print(f"Schedule items: {total_schedule}")
    print(f"Habit logs: {total_habits}")
    print(f"Study period: Jul 31 - Nov 28, 2026")


if __name__ == "__main__":
    import sys
    import asyncio
    asyncio.run(init_db())
    if len(sys.argv) > 1:
        sched_path = sys.argv[1]
        course_path = sys.argv[2] if len(sys.argv) > 2 else sched_path
        import_data(sched_path, course_path)
    else:
        print("Database tables created. Usage: python database.py <schedule.xlsx> [courses.xlsx]")
